#!/usr/bin/env python3
"""
Comprehensive Product Functionality Test Script
Tests: Add Product, Edit Product, Bulk Upload, Add to Cart
"""

import requests
import json
import time
import os
from datetime import datetime

# Configuration
BASE_URL = "http://localhost:5000"
API_BASE = f"{BASE_URL}/api"

# Test credentials
TEST_USERS = {
    "manufacturer": {
        "email": "m@demo.com",
        "password": "Demo@123"
    },
    "distributor": {
        "email": "d@demo.com", 
        "password": "Demo@123"
    },
    "retailer": {
        "email": "r@demo.com",
        "password": "Demo@123"
    }
}

def make_request(method, endpoint, data=None, token=None, is_form_data=False):
    """Make API request with proper headers"""
    url = f"{API_BASE}{endpoint}"
    headers = {}
    
    if token:
        headers['Authorization'] = f'Bearer {token}'
    
    if data and not is_form_data:
        headers['Content-Type'] = 'application/json'
    
    try:
        if method == "GET":
            response = requests.get(url, headers=headers)
        elif method == "POST":
            if is_form_data:
                response = requests.post(url, files=data, headers=headers)
            else:
                response = requests.post(url, json=data, headers=headers)
        elif method == "PUT":
            response = requests.put(url, json=data, headers=headers)
        elif method == "DELETE":
            response = requests.delete(url, headers=headers)
        else:
            raise ValueError(f"Unsupported method: {method}")
        
        return response
    except requests.exceptions.RequestException as e:
        print(f"❌ Request failed: {e}")
        return None

def login_user(email, password):
    """Login user and return token"""
    print(f"🔐 Logging in {email}...")
    
    login_data = {
        "email": email,
        "password": password
    }
    
    response = make_request("POST", "/auth/login", login_data)
    
    if response and response.status_code == 200:
        result = response.json()
        token = result.get('access_token')
        user = result.get('user')
        print(f"✅ Login successful for {user.get('role')}: {user.get('firstName')} {user.get('lastName')}")
        return token, user
    else:
        print(f"❌ Login failed for {email}")
        if response:
            print(f"   Status: {response.status_code}")
            print(f"   Response: {response.text}")
        return None, None

def test_add_product(token, user_role):
    """Test adding a new product"""
    print(f"\n📦 Testing Add Product for {user_role}...")
    
    product_data = {
        "name": f"Test Product {user_role}",
        "description": f"This is a test product for {user_role}",
        "sku": f"TEST-{user_role.upper()}-{int(time.time())}",
        "basePrice": 99.99,
        "imageUrl": "https://via.placeholder.com/300x200",
        "brand": "Test Brand",
        "unit": "pcs"
    }
    
    response = make_request("POST", "/products", product_data, token)
    
    if response and response.status_code == 201:
        product = response.json()
        print(f"✅ Product added successfully!")
        print(f"   ID: {product.get('id')}")
        print(f"   Name: {product.get('name')}")
        print(f"   SKU: {product.get('sku')}")
        return product
    else:
        print(f"❌ Failed to add product")
        if response:
            print(f"   Status: {response.status_code}")
            print(f"   Response: {response.text}")
        return None

def test_edit_product(token, product_id, user_role):
    """Test editing a product"""
    print(f"\n✏️ Testing Edit Product for {user_role}...")
    
    edit_data = {
        "name": f"Updated Test Product {user_role}",
        "description": f"This is an updated test product for {user_role}",
        "basePrice": 149.99,
        "imageUrl": "https://via.placeholder.com/400x300"
    }
    
    response = make_request("PUT", f"/products/{product_id}", edit_data, token)
    
    if response and response.status_code == 200:
        product = response.json()
        print(f"✅ Product updated successfully!")
        print(f"   New Name: {product.get('name')}")
        print(f"   New Price: {product.get('basePrice')}")
        return product
    else:
        print(f"❌ Failed to update product")
        if response:
            print(f"   Status: {response.status_code}")
            print(f"   Response: {response.text}")
        return None

def test_bulk_upload(token, user_role):
    """Test bulk upload functionality"""
    print(f"\n📤 Testing Bulk Upload for {user_role}...")
    
    # Create a simple Excel file for testing
    import openpyxl
    from io import BytesIO
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    
    # Add headers
    headers = ['name', 'description', 'sku', 'base_price', 'brand']
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Add sample data
    sample_data = [
        ['Bulk Product 1', 'First bulk product', f'BULK-{user_role.upper()}-001', 25.99, 'Bulk Brand'],
        ['Bulk Product 2', 'Second bulk product', f'BULK-{user_role.upper()}-002', 35.99, 'Bulk Brand'],
        ['Bulk Product 3', 'Third bulk product', f'BULK-{user_role.upper()}-003', 45.99, 'Bulk Brand'],
    ]
    
    for row, data in enumerate(sample_data, 2):
        for col, value in enumerate(data, 1):
            ws.cell(row=row, column=col, value=value)
    
    # Save to BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    # Prepare form data
    files = {'file': ('bulk_products.xlsx', excel_file, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
    
    response = make_request("POST", "/products/bulk-upload", files, token, is_form_data=True)
    
    if response and response.status_code == 200:
        result = response.json()
        print(f"✅ Bulk upload completed!")
        print(f"   Added: {result.get('added')}")
        print(f"   Failed: {result.get('failed')}")
        if result.get('errors'):
            print(f"   Errors: {result.get('errors')}")
        return result
    else:
        print(f"❌ Bulk upload failed")
        if response:
            print(f"   Status: {response.status_code}")
            print(f"   Response: {response.text}")
        return None

def test_add_to_cart(token, product_id, user_role):
    """Test adding product to cart"""
    print(f"\n🛒 Testing Add to Cart for {user_role}...")
    
    cart_data = {
        "productId": product_id,
        "quantity": 2
    }
    
    response = make_request("POST", "/cart", cart_data, token)
    
    if response and response.status_code == 200:
        result = response.json()
        print(f"✅ Product added to cart successfully!")
        print(f"   Message: {result.get('message')}")
        return True
    else:
        print(f"❌ Failed to add product to cart")
        if response:
            print(f"   Status: {response.status_code}")
            print(f"   Response: {response.text}")
        return False

def test_get_cart(token, user_role):
    """Test getting cart contents"""
    print(f"\n🛒 Testing Get Cart for {user_role}...")
    
    response = make_request("GET", "/cart", token=token)
    
    if response and response.status_code == 200:
        cart = response.json()
        print(f"✅ Cart retrieved successfully!")
        print(f"   Total Items: {cart.get('totalItems')}")
        print(f"   Total Amount: {cart.get('totalAmount')}")
        if cart.get('items'):
            for item in cart['items']:
                print(f"   - {item.get('productName')} x{item.get('quantity')} @ ${item.get('unitPrice')}")
        return cart
    else:
        print(f"❌ Failed to get cart")
        if response:
            print(f"   Status: {response.status_code}")
            print(f"   Response: {response.text}")
        return None

def test_get_products(token, user_role):
    """Test getting products"""
    print(f"\n📋 Testing Get Products for {user_role}...")
    
    response = make_request("GET", "/products", token=token)
    
    if response and response.status_code == 200:
        products = response.json()
        print(f"✅ Products retrieved successfully!")
        print(f"   Total Products: {len(products)}")
        for product in products[:3]:  # Show first 3 products
            print(f"   - {product.get('name')} (SKU: {product.get('sku')})")
        return products
    else:
        print(f"❌ Failed to get products")
        if response:
            print(f"   Status: {response.status_code}")
            print(f"   Response: {response.text}")
        return None

def main():
    """Run comprehensive product functionality tests"""
    print("🚀 Starting Comprehensive Product Functionality Tests")
    print("=" * 60)
    
    results = {}
    
    for role, credentials in TEST_USERS.items():
        print(f"\n{'='*20} Testing {role.upper()} {'='*20}")
        
        # Login
        token, user = login_user(credentials['email'], credentials['password'])
        if not token:
            print(f"❌ Skipping {role} tests due to login failure")
            continue
        
        results[role] = {
            'login': True,
            'add_product': False,
            'edit_product': False,
            'bulk_upload': False,
            'add_to_cart': False,
            'get_cart': False,
            'get_products': False
        }
        
        # Test 1: Get Products
        products = test_get_products(token, role)
        if products:
            results[role]['get_products'] = True
        
        # Test 2: Add Product (only for manufacturer and distributor)
        if role in ['manufacturer', 'distributor']:
            product = test_add_product(token, role)
            if product:
                results[role]['add_product'] = True
                product_id = product['id']
                
                # Test 3: Edit Product
                updated_product = test_edit_product(token, product_id, role)
                if updated_product:
                    results[role]['edit_product'] = True
                
                # Test 4: Add to Cart (using the created product)
                cart_success = test_add_to_cart(token, product_id, role)
                if cart_success:
                    results[role]['add_to_cart'] = True
            else:
                product_id = None
        else:
            # For retailers, try to add an existing product to cart
            if products:
                product_id = products[0]['id']
                cart_success = test_add_to_cart(token, product_id, role)
                if cart_success:
                    results[role]['add_to_cart'] = True
        
        # Test 5: Bulk Upload (only for manufacturer and distributor)
        if role in ['manufacturer', 'distributor']:
            bulk_result = test_bulk_upload(token, role)
            if bulk_result:
                results[role]['bulk_upload'] = True
        
        # Test 6: Get Cart
        cart = test_get_cart(token, role)
        if cart:
            results[role]['get_cart'] = True
        
        print(f"\n✅ {role.upper()} Tests Completed")
    
    # Print Summary
    print(f"\n{'='*60}")
    print("📊 TEST SUMMARY")
    print("=" * 60)
    
    for role, result in results.items():
        print(f"\n{role.upper()}:")
        for test, passed in result.items():
            status = "✅" if passed else "❌"
            print(f"  {status} {test.replace('_', ' ').title()}")
    
    print(f"\n🎉 All tests completed at {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

if __name__ == "__main__":
    main()
